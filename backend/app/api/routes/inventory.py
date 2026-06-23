"""
Inventory routes:
  GET  /api/inventory                — list all inventory
  GET  /api/inventory/check          — availability check
  GET  /api/inventory/forecast       — stockout forecast (Feature 2)
  POST /api/inventory/import-excel   — bulk import from .xlsx (Feature 4)
  GET  /api/inventory/export-excel   — download inventory as .xlsx (Feature 4)
"""
from __future__ import annotations

import io
from typing import Any

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app import crud
from app.database import get_db
from app.models import Inventory
from app.schemas import InventoryAvailability, InventoryRead, StockoutForecast
from app.services.forecast_service import compute_forecasts

router = APIRouter(prefix="/inventory", tags=["Inventory"])


# ── Existing endpoints (unchanged) ──────────────────────────────────────────

@router.get("", response_model=list[InventoryRead])
def list_inventory(db: Session = Depends(get_db)):
    return crud.get_inventory(db)


@router.get("/check", response_model=InventoryAvailability)
def check_inventory(lens_type: str, power: float, db: Session = Depends(get_db)):
    return crud.get_inventory_availability(db, lens_type, power)


# ── Feature 2: Stockout Forecast ────────────────────────────────────────────

@router.get("/forecast", response_model=list[StockoutForecast])
def stockout_forecast(db: Session = Depends(get_db)):
    """
    Returns stockout forecasts for all SKUs, sorted by urgency (soonest first).
    Uses historical order consumption rate (last 60 days) with linear regression.
    """
    return compute_forecasts(db)


# ── Feature 4: Excel Import ─────────────────────────────────────────────────

EXPECTED_COLUMNS = {"sku", "name", "stock", "reorder_threshold"}
# The name column maps to lens_type in our schema.
# Column aliases accepted (case-insensitive):
COLUMN_MAP = {
    "sku": "sku",
    "name": "name",
    "lens_type": "name",        # allow "lens_type" as alias for "name"
    "stock": "stock",
    "quantity": "stock",        # allow "quantity" as alias for "stock"
    "reorder_threshold": "reorder_threshold",
    "reorder_level": "reorder_threshold",  # allow "reorder_level" as alias
    "power": "power",
}


@router.post("/import-excel")
async def import_inventory_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """
    Accepts an .xlsx file with columns: SKU (=id or ignored), name (lens_type),
    stock (quantity), reorder_threshold (reorder_level), and optionally power.

    Returns row-level validation errors instead of failing the whole import.
    """
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=415, detail="Only .xlsx files are supported")

    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
        ws = wb.active
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Could not read Excel file: {exc}") from exc

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Excel file is empty")

    # Parse header row
    raw_headers = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
    mapped_headers = [COLUMN_MAP.get(h, h) for h in raw_headers]

    imported = 0
    errors: list[dict] = []

    for row_idx, row in enumerate(rows[1:], start=2):
        row_data: dict[str, Any] = dict(zip(mapped_headers, row))

        # Validate required fields
        lens_type = str(row_data.get("name", "") or "").strip()
        if not lens_type:
            errors.append({"row": row_idx, "error": "Missing lens type (name column)"})
            continue

        try:
            stock = int(row_data.get("stock") or 0)
        except (TypeError, ValueError):
            errors.append({"row": row_idx, "error": f"Invalid stock value: {row_data.get('stock')!r}"})
            continue

        if stock < 0:
            errors.append({"row": row_idx, "error": f"Stock cannot be negative (got {stock})"})
            continue

        try:
            reorder = int(row_data.get("reorder_threshold") or 20)
        except (TypeError, ValueError):
            reorder = 20

        try:
            power = float(row_data.get("power") or 0.0)
        except (TypeError, ValueError):
            power = 0.0

        # Upsert: update existing row (matched by lens_type + power) or create new
        existing = db.query(Inventory).filter(
            Inventory.lens_type == lens_type,
            Inventory.power == power,
        ).first()

        if existing:
            existing.quantity = stock
            existing.reorder_level = reorder
            db.add(existing)
        else:
            db.add(Inventory(
                lens_type=lens_type,
                power=power,
                quantity=stock,
                reorder_level=reorder,
            ))

        imported += 1

    if imported > 0:
        db.commit()

    return {
        "imported": imported,
        "errors": errors,
        "total_rows": len(rows) - 1,
    }


# ── Feature 4: Excel Export ─────────────────────────────────────────────────

@router.get("/export-excel")
def export_inventory_excel(db: Session = Depends(get_db)):
    """
    Returns current inventory as a downloadable .xlsx file.
    Columns: ID, Lens Type, Power, Quantity, Reorder Level, Last Updated.
    """
    items = crud.get_inventory(db)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory"

    headers = ["ID", "Lens Type", "Power", "Quantity (Stock)", "Reorder Level", "Last Updated"]
    ws.append(headers)

    # Style header row
    from openpyxl.styles import Font, PatternFill
    header_fill = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for item in items:
        ws.append([
            item.id,
            item.lens_type,
            item.power,
            item.quantity,
            item.reorder_level,
            item.updated_at.strftime("%Y-%m-%d %H:%M") if item.updated_at else "",
        ])

    # Auto-fit column widths (approximate)
    col_widths = [6, 20, 10, 18, 16, 20]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=inventory.xlsx"},
    )
